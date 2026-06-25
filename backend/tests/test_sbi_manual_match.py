"""
Tests for SBI Kiosk manual match (persistent overlay).

P01–P04 results are delete-and-recreated on every run (#17), so a manual match is
stored keyed by the row's stable business key and overlaid at READ time. The key
property: a manual match SURVIVES a re-run (which gives the row a new id).
"""
import openpyxl
import pytest
from fastapi import HTTPException
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from models.database import Base, User, SBIP02Result, SBIManualMatch
from routes.sbi_kiosk import (
    create_manual_match, delete_manual_match, list_manual_matches,
    get_p02_results, _build_sbi_export, ManualMatchIn,
)

DATE = "2026-06-20"
USER = User(id="u1", username="raj", role="user", permissions='{"src_assign": true}')


@pytest.fixture
def db():
    engine = create_engine("sqlite:///:memory:", connect_args={"check_same_thread": False})
    Base.metadata.create_all(engine)
    Session = sessionmaker(bind=engine, autoflush=False, autocommit=False)
    s = Session()
    yield s
    s.close()


def _p02(db, ref, status="Unmatched_Bank", bank_type="DR"):
    r = SBIP02Result(recon_date=DATE, reference_number=ref, bank_type=bank_type,
                     match_status=status, bank_amount=100)
    db.add(r); db.commit(); return r


def _p02_rows(db):
    return get_p02_results(recon_date=DATE, match_status=None, page=1, page_size=100,
                           db=db, current_user=USER)


def test_manual_match_overlays_status(db):
    r = _p02(db, "REF1")
    out = create_manual_match(ManualMatchIn(process="p02", result_id=r.id,
                              remark="known typo in ref"), db=db, current_user=USER)
    assert out["match_status"] == "Manual_Matched"
    res = _p02_rows(db)
    assert res["manual_matched_count"] == 1
    row = next(x for x in res["rows"] if x["reference_number"] == "REF1")
    assert row["match_status"] == "Manual_Matched"
    assert row["manual_remark"] == "known typo in ref"


def test_manual_match_survives_a_rerun(db):
    # This is the whole point: delete-and-recreate the result row (new id) and the
    # manual match must still apply, keyed on the stable business key.
    r = _p02(db, "REF1")
    create_manual_match(ManualMatchIn(process="p02", result_id=r.id, remark="manual link ok"),
                        db=db, current_user=USER)
    db.delete(r); db.commit()                      # simulate the P02 re-run wiping results
    _p02(db, "REF1")                               # recreated with a NEW id
    row = next(x for x in _p02_rows(db)["rows"] if x["reference_number"] == "REF1")
    assert row["match_status"] == "Manual_Matched"  # survived


def test_remark_required(db):
    r = _p02(db, "REF2")
    with pytest.raises(HTTPException) as e:
        create_manual_match(ManualMatchIn(process="p02", result_id=r.id, remark="x"),
                            db=db, current_user=USER)
    assert e.value.status_code == 400


def test_delete_reverts_to_algorithm_status(db):
    r = _p02(db, "REF3")
    out = create_manual_match(ManualMatchIn(process="p02", result_id=r.id, remark="manual link ok"),
                              db=db, current_user=USER)
    delete_manual_match(out["id"], db=db, current_user=USER)
    res = _p02_rows(db)
    row = next(x for x in res["rows"] if x["reference_number"] == "REF3")
    assert row["match_status"] == "Unmatched_Bank"   # reverted to the run output
    assert res["manual_matched_count"] == 0


def test_export_reflects_manual_match(db):
    r = _p02(db, "REF4")
    create_manual_match(ManualMatchIn(process="p02", result_id=r.id, remark="export check ok"),
                        db=db, current_user=USER)
    wb = openpyxl.load_workbook(_build_sbi_export(db, ["p02"], recon_date=DATE))
    ws = wb["Bank vs Transaction"]
    header = [c.value for c in ws[1]]
    sidx = header.index("match_status") + 1
    statuses = {ws.cell(row=row, column=sidx).value for row in range(2, ws.max_row + 1)}
    assert "Manual_Matched" in statuses


def test_p01_is_not_a_supported_process(db):
    with pytest.raises(HTTPException) as e:
        create_manual_match(ManualMatchIn(process="p01", result_id="x", remark="not allowed here"),
                            db=db, current_user=USER)
    assert e.value.status_code == 400
