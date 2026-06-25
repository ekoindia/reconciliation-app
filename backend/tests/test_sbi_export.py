"""
Tests for the SBI Kiosk recon export (fixes the blank-export bug).

SBI reconciles in its own sbi_p0x_results tables, not the core `transactions`
ledger, so the generic /api/reports export returned blank for partner=kiosk.
_build_sbi_export reads the real result tables and always produces a valid,
openpyxl-readable workbook — one sheet per process, with a header row even when
there is no data.
"""
import openpyxl
import pytest
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from models.database import Base, SBIP01Result, SBIP02Result
from routes.sbi_kiosk import _build_sbi_export, _SBI_EXPORTS

DATE = "2026-06-20"


@pytest.fixture
def db():
    engine = create_engine("sqlite:///:memory:", connect_args={"check_same_thread": False})
    Base.metadata.create_all(engine)
    Session = sessionmaker(bind=engine, autoflush=False, autocommit=False)
    s = Session()
    yield s
    s.close()


def _load(buf):
    return openpyxl.load_workbook(buf)


def test_export_all_has_one_sheet_per_process(db):
    wb = _load(_build_sbi_export(db, list(_SBI_EXPORTS)))
    assert wb.sheetnames == ["Bank vs Settlement", "Bank vs Transaction",
                             "CSP-Txn-Bank", "Wallet Balance"]


def test_export_includes_real_p01_data(db):
    db.add(SBIP01Result(recon_date=DATE, ko_id="KO1", wallet_withdrawn=500.0,
                        bank_settled=500.0, difference=0.0, status="CREDITED"))
    db.add(SBIP01Result(recon_date=DATE, ko_id="KO2", wallet_withdrawn=300.0,
                        bank_settled=0.0, difference=300.0, status="PENDING"))
    db.commit()
    ws = _load(_build_sbi_export(db, ["p01"], recon_date=DATE))["Bank vs Settlement"]
    # header + 2 rows
    assert ws.max_row == 3
    header = [c.value for c in ws[1]]
    assert "ko_id" in header and "bank_settled" in header and "status" in header
    ko_ids = {ws.cell(row=r, column=header.index("ko_id") + 1).value for r in (2, 3)}
    assert ko_ids == {"KO1", "KO2"}


def test_empty_export_is_still_a_valid_file_with_headers(db):
    # The bug was a BLANK file; an empty result set must still yield a header-only
    # sheet that opens cleanly, not an empty/blank download.
    ws = _load(_build_sbi_export(db, ["p02"], recon_date="2099-01-01"))["Bank vs Transaction"]
    assert ws.max_row == 1                          # header only, no data
    assert [c.value for c in ws[1]][:2] == ["recon_date", "reference_number"]


def test_match_status_filter_applies_only_where_present(db):
    db.add(SBIP02Result(recon_date=DATE, reference_number="R1", match_status="Matched", bank_amount=10))
    db.add(SBIP02Result(recon_date=DATE, reference_number="R2", match_status="Unmatched_Bank", bank_amount=20))
    db.commit()
    ws = _load(_build_sbi_export(db, ["p02"], recon_date=DATE, match_status="Matched"))["Bank vs Transaction"]
    assert ws.max_row == 2                           # header + the one Matched row
