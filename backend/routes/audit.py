"""
Audit Log API
Read-only endpoint — browse activity history with filters.
"""
import json
from fastapi import APIRouter, Depends, Query
from sqlalchemy.orm import Session
from typing import Optional
from models.database import get_db, AuditLog, User
from core.auth import require_permission

import datetime as _dt
_IST = _dt.timezone(_dt.timedelta(hours=5, minutes=30))

def _ist_str(dt):
    """Render a stored (naive UTC) datetime as IST 'YYYY-MM-DD HH:MM:SS' for export.
    The UI converts UTC→IST in the API layer; Excel exports bypass that, so we
    convert here."""
    if not dt:
        return ""
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=_dt.timezone.utc)
    return dt.astimezone(_IST).strftime("%Y-%m-%d %H:%M:%S")

router = APIRouter(prefix="/api/audit", tags=["audit"])


def _safe_json(raw):
    """Parse a detail/previous_state value as JSON, falling back to the raw string.
    Some writers (e.g. early E-Value routes) stored a plain string or a Python
    dict repr; this keeps the audit endpoint from 500-ing on those rows."""
    if not raw:
        return None
    try:
        return json.loads(raw)
    except Exception:
        return str(raw)


# Actions that are always human-triggered (used to infer type for legacy NULL rows)
HUMAN_ACTIONS = {
    "upload", "clear", "delete_selected", "run_recon", "run_neft_d1",
    "run_internal_match", "interbank_match", "manual_match", "assign_src",
    "bulk_src", "tag_adhoc", "human_override", "override_status",
    "unmatch", "rematch", "bulk_override", "manual_interbank_match",
    "rematch_break_other", "override_break_match_other", "run_range",
}


def _resolve_type(log) -> str:
    """
    Effective action_type for a log row.
    - New rows (action_type set explicitly) → use it directly.
    - Legacy rows (action_type NULL) → infer:
        * username 'system' or action 'auto_upload' → app
        * action in HUMAN_ACTIONS with a real username   → human
        * everything else                                → app
    """
    at = getattr(log, "action_type", None)
    if at:
        return at
    if log.username in ("system", "auto_upload", None):
        return "app"
    if log.action in HUMAN_ACTIONS:
        return "human"
    return "app"


@router.get("/logs")
def list_audit_logs(
    action:      Optional[str] = None,
    action_type: Optional[str] = None,   # "app" | "human" | "" (all)
    user_id:     Optional[str] = None,
    entity_type: Optional[str] = None,
    from_date:   Optional[str] = None,   # YYYY-MM-DD
    to_date:     Optional[str] = None,
    page:        int = Query(1, ge=1),
    page_size:   int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("audit_read")),
):
    q = db.query(AuditLog)

    if action:      q = q.filter(AuditLog.action == action)
    if user_id:     q = q.filter(AuditLog.user_id == user_id)
    if entity_type: q = q.filter(AuditLog.entity_type == entity_type)
    if from_date:   q = q.filter(AuditLog.created_at >= f"{from_date} 00:00:00")
    if to_date:     q = q.filter(AuditLog.created_at <= f"{to_date} 23:59:59")

    # action_type filter — NULL rows in DB are legacy; treat as "app"
    if action_type == "human":
        q = q.filter(AuditLog.action_type == "human")
    elif action_type == "app":
        from sqlalchemy import or_
        q = q.filter(
            or_(AuditLog.action_type == "app", AuditLog.action_type.is_(None))
        )
    # "" or None → no filter (show all)

    total = q.count()
    logs  = q.order_by(AuditLog.created_at.desc()) \
             .offset((page - 1) * page_size).limit(page_size).all()

    return {
        "total":     total,
        "page":      page,
        "page_size": page_size,
        "logs": [
            {
                "id":             log.id,
                "username":       log.username,
                "action":         log.action,
                "action_type":    _resolve_type(log),
                "entity_type":    log.entity_type,
                "entity_id":      log.entity_id,
                "detail":         _safe_json(log.detail),
                "previous_state": _safe_json(getattr(log, "previous_state", None)),
                "created_at":     str(log.created_at),
            }
            for log in logs
        ],
    }


@router.get("/actions")
def list_action_types(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("audit_read")),
):
    """Distinct action names in the log — for the filter dropdown."""
    rows = db.query(AuditLog.action).distinct().all()
    return sorted([r[0] for r in rows if r[0]])


@router.get("/logs/export")
def export_audit_logs(
    action:      Optional[str] = None,
    action_type: Optional[str] = None,
    from_date:   Optional[str] = None,
    to_date:     Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("audit_read")),
):
    """Export audit log to Excel — same filters as /logs."""
    import io, pandas as pd
    from fastapi.responses import StreamingResponse

    q = db.query(AuditLog)
    if action:      q = q.filter(AuditLog.action == action)
    if from_date:   q = q.filter(AuditLog.created_at >= f"{from_date} 00:00:00")
    if to_date:     q = q.filter(AuditLog.created_at <= f"{to_date} 23:59:59")
    if action_type == "human":
        q = q.filter(AuditLog.action_type == "human")
    elif action_type == "app":
        from sqlalchemy import or_
        q = q.filter(or_(AuditLog.action_type == "app", AuditLog.action_type.is_(None)))

    logs = q.order_by(AuditLog.created_at.desc()).limit(10000).all()

    rows = [{
        "Timestamp":   _ist_str(log.created_at),
        "Actor Type":  _resolve_type(log),
        "User":        log.username or "",
        "Action":      log.action or "",
        "Entity Type": log.entity_type or "",
        "Entity ID":   (log.entity_id or "")[:16],
        "Detail":      (json.dumps(_safe_json(log.detail)) if log.detail else ""),
    } for log in logs]

    df = pd.DataFrame(rows)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Audit Log")
        ws = writer.sheets["Audit Log"]
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = max(
                len(str(col[0].value or "")), max((len(str(c.value or "")) for c in col), default=10)
            ) + 2
        from openpyxl.styles import PatternFill, Font as XFont
        for cell in ws[1]:
            cell.fill = PatternFill("solid", fgColor="1E3A5F")
            cell.font = XFont(bold=True, color="FFFFFF")
    output.seek(0)
    return StreamingResponse(output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=audit_log.xlsx"})
